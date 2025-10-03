import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine
from config import DB_CONFIG
import os

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# DB CONNECTION
# -----------------------------
engine = create_engine(
    f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
)

# Make sure folders exist
(os.makedirs("charts", exist_ok=True))
(os.makedirs("exports", exist_ok=True))

# -----------------------------
# Helper: Run SQL → DataFrame
# -----------------------------
def run_query(sql):
    return pd.read_sql(sql, engine)

# -----------------------------
# 1. VISUALIZATIONS (6 charts)
# -----------------------------
def generate_charts():
    queries = {
        "pie": {
            "sql": """
                SELECT c.category_name, COUNT(p.product_id) AS product_count
                FROM production.products p
                JOIN production.categories c ON p.category_id = c.category_id
                GROUP BY c.category_name;
            """,
            "title": "Distribution of Products by Category",
            "kind": "pie"
        },
        "bar": {
            "sql": """
                SELECT s.store_name, SUM(oi.quantity * p.list_price) AS revenue
                FROM sales.order_items oi
                JOIN production.products p ON oi.product_id = p.product_id
                JOIN sales.orders o ON o.order_id = oi.order_id
                JOIN sales.stores s ON o.store_id = s.store_id
                GROUP BY s.store_name
                ORDER BY revenue DESC;
            """,
            "title": "Revenue by Store",
            "kind": "bar"
        },
        "barh": {
            "sql": """
                SELECT b.brand_name, AVG(p.list_price) AS avg_price
                FROM production.products p
                JOIN production.brands b ON p.brand_id = b.brand_id
                GROUP BY b.brand_name
                ORDER BY avg_price DESC;
            """,
            "title": "Average Product Price by Brand",
            "kind": "barh"
        },
        "line": {
            "sql": """
                SELECT DATE(o.order_date) AS order_date, SUM(oi.quantity * p.list_price) AS daily_sales
                FROM sales.orders o
                JOIN sales.order_items oi ON o.order_id = oi.order_id
                JOIN production.products p ON oi.product_id = p.product_id
                GROUP BY DATE(o.order_date)
                ORDER BY order_date;
            """,
            "title": "Daily Sales Trend",
            "kind": "line"
        },
        "hist": {
            "sql": """
                SELECT p.list_price
                FROM production.products p;
            """,
            "title": "Distribution of Product Prices",
            "kind": "hist"
        },
        "scatter": {
            "sql": """
                SELECT c.city, COUNT(o.order_id) AS num_orders, SUM(oi.quantity * p.list_price) AS revenue
                FROM sales.customers c
                JOIN sales.orders o ON c.customer_id = o.customer_id
                JOIN sales.order_items oi ON o.order_id = oi.order_id
                JOIN production.products p ON oi.product_id = p.product_id
                GROUP BY c.city
                ORDER BY revenue DESC;
            """,
            "title": "Orders vs Revenue by City",
            "kind": "scatter"
        }
    }

    for chart_type, details in queries.items():
        df = run_query(details["sql"])

        plt.figure(figsize=(8, 6))

        if details["kind"] == "pie":
            df.set_index("category_name")["product_count"].plot.pie(autopct="%1.1f%%")
            plt.ylabel("")
        
        elif details["kind"] == "bar":
            df.plot(kind="bar", x="store_name", y="revenue", legend=False)
            plt.ylabel("Revenue")
        
        elif details["kind"] == "barh":
            df.plot(kind="barh", x="brand_name", y="avg_price", legend=False)
            plt.xlabel("Average Price")
        
        elif details["kind"] == "line":
            df.plot(kind="line", x="order_date", y="daily_sales", marker="o", legend=False)
            plt.ylabel("Sales")
        
        elif details["kind"] == "hist":
            df["list_price"].plot.hist(bins=20)
            plt.xlabel("Price")
        
        elif details["kind"] == "scatter":
            df.plot(kind="scatter", x="num_orders", y="revenue")
            plt.xlabel("Number of Orders")
            plt.ylabel("Revenue")

        plt.title(details["title"])
        filename = f"charts/{chart_type}.png"
        plt.savefig(filename)
        plt.close()

        print(f"✅ {chart_type} chart saved → {filename}, rows={len(df)}")

# -----------------------------
# 2. TIME SLIDER (PLOTLY)
# -----------------------------
def interactive_plot():
    sql = """
        SELECT DATE(o.order_date) AS order_date, s.store_name, SUM(oi.quantity) AS qty
        FROM sales.orders o
        JOIN sales.order_items oi ON o.order_id = oi.order_id
        JOIN production.products p ON oi.product_id = p.product_id
        JOIN sales.stores s ON o.store_id = s.store_id
        GROUP BY DATE(o.order_date), s.store_name
        ORDER BY order_date;
    """
    df = run_query(sql)
    df["order_date"] = df["order_date"].astype(str)

    fig = px.bar(
        df,
        x="store_name",
        y="qty",
        animation_frame="order_date",
        color="store_name",
        title="Store Sales Over Time"
    )

    fig.update_layout(
        xaxis_title="Store",
        yaxis_title="Quantity Sold",
        legend_title="Store"
    )

    fig.show()

# -----------------------------
# 3. EXPORT TO EXCEL (MODIFIED)
# -----------------------------
def export_to_excel(dataframes_dict, filename):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet, df in dataframes_dict.items():

            # ✅ Sort OrdersReport by order_revenue ascending (least to most)
            if sheet == "OrdersReport" and "order_revenue" in df.columns:
                df = df.sort_values(by="order_revenue", ascending=True)

            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # ✅ Only apply gradient to order_revenue column in OrdersReport
        if sheet == "OrdersReport":
            for col in ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=1):
                if col[0].value == "order_revenue":
                    col_letter = col[0].column_letter
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFCCCC",
                        end_type="max", end_color="00FF00"
                    )
                    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
                    break

    wb.save(filename)
    print(f"✅ Created file {filename}, {len(dataframes_dict)} sheets, rows ~{ws.max_row}")


# -----------------------------
# MAIN
# -----------------------------
if __name__ == "__main__":
    generate_charts()
    interactive_plot()

    dfs = {
        "OrdersReport": run_query("""
            SELECT o.order_id, o.order_date, c.first_name || ' ' || c.last_name AS customer,
                   s.store_name, SUM(oi.quantity * p.list_price) AS order_revenue
            FROM sales.orders o
            JOIN sales.customers c ON o.customer_id = c.customer_id
            JOIN sales.stores s ON o.store_id = s.store_id
            JOIN sales.order_items oi ON o.order_id = oi.order_id
            JOIN production.products p ON oi.product_id = p.product_id
            GROUP BY o.order_id, o.order_date, c.first_name, c.last_name, s.store_name
            ORDER BY o.order_date;
        """),
        "ProductsStock": run_query("""
            SELECT p.product_name, b.brand_name, c.category_name, p.list_price, st.store_id, st.quantity
            FROM production.products p
            JOIN production.brands b ON p.brand_id = b.brand_id
            JOIN production.categories c ON p.category_id = c.category_id
            JOIN production.stocks st ON p.product_id = st.product_id
            ORDER BY b.brand_name, p.product_name;
        """)
    }

    export_to_excel(dfs, "exports/report.xlsx")
